from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def classifica_pfpj(pam1):

    wb = load_workbook('excel files/Importação de processos (Aberto) Corrigido.xlsx')

    ws = wb.active

    redFill = PatternFill(start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid')

    for cell in ws['{}'.format(pam1)]:
        if 'S/A' in cell.value or 'S/a' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'LTDA' in cell.value or 'Ltda' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif ' CIA' in cell.value or ' Cia' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'COMPANH' in cell.value or 'Companh' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'CONDOM' in cell.value or 'Condom' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'ASSOCIA' in cell.value or 'Associa' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'SEGUR' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif '-' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'INSTI' in cell.value or 'Insti' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)                  
        elif 'ADVO' in cell.value or 'Advo' in cell.value:
            a = cell.row
            b = 14
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)

    wb.save("excel files/importacao_format.xlsx")