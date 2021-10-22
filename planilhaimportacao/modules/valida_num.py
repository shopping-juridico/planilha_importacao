import re
from openpyxl.styles import PatternFill

def so_numeros(num):
    return re.sub('[^0-9]', '', num)

def conta_num(col_num_ini: str, col_num_ant: int, ws):

    yellowFill = PatternFill(start_color='FFC7CE',
                end_color='FFC7CE',
                fill_type='solid')

    for cell in ws[f'{col_num_ini}']:
        if(cell.value is not None):
            if cell.value == "Número CNJ":
                continue
            if len(so_numeros(cell.value)) != 20:
                # cell.fill = yellowFill
                copy = cell.value
                paste = ws.cell(row=cell.row, column=col_num_ant)
                paste.value = copy
                cell.value = ""
                print("Classificado: Numeração antiga")
    return ws
            
        