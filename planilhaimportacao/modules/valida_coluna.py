from openpyxl import load_workbook
from modules.validacpf import ValidaCpf
from modules.validacnpj import ValidaCnpj
from openpyxl.styles import PatternFill

def valida_coluna(pam2):
    
    wb = load_workbook('excel files/Importação de processos (Aberto) Corrigido.xlsx')

    ws = wb.active

    redFill = PatternFill(start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid')
        
    for cell in ws['{}'.format(pam2)]:
        if(cell.value is not None):
            if cell.value == "Contrário Principal - CPF/CNPJ":
                continue
            if cell.value == "Cliente principal - CPF/CNPJ":
                continue
            cpf = ValidaCpf(cell.value)
            cnpj = ValidaCnpj(cell.value)
            if cpf.valida():
                #a = 2
                print('CPF válido')
                #print(cell.value)
            elif cnpj.valida():
                #a = 3
                print('CNPJ válido')
                #print(cell.value)
            else:
                cell.fill = redFill
                print('CPF inválido')
        if cell.value is None:
            cell.fill = redFill
            print("Campo vazio")

    wb.save("excel files/importacao_format.xlsx")
