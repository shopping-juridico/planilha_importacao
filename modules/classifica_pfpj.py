from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def classifica_pfpj(pam1):

    wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

    ws = wb.active

    redFill = PatternFill(start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid')

    for cell in ws['{}'.format(pam1)]:
        if 'S/A' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'LTDA' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'CIA' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'COMPANH' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'COND' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'ASSOCI' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'SEGUR' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif '-' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'INSTI' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)                  
        elif 'ADVO' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)

    wb.save("/home/vitor/projetos/planilha_importacao/importacao_format.xlsx")