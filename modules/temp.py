from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from validacpf import ValidaCpf
from validacnpj import ValidaCnpj

wb = load_workbook('excel files/Importação de processos (Aberto) Corrigido.xlsx')

ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

for cell in ws['R']:
    v = [cell.value, 'SEGUROS']
    if 'SEGUROS' in v:
        a = cell.row
        b = 20
        print(a, b)
        c = ws.cell(row=a, column=b)
        c.value = 'PJ'
        print(c.value)
        
wb.save("excel files/importacao_format.xlsx")
