from openpyxl import load_workbook

def fatia(num_cnj):
    #num_cnj = "0000100-15.2008.816.0000"
    #fatia_um = num_cnj[:7]
    #fatia_dois = num_cnj[8:10]
    #fatia_tres = num_cnj[11:15]
    fatia_quatro = num_cnj[16:17]
    #fatia_cinco = num_cnj[20:24]
    return fatia_quatro

def depara_orgao(pam3):
    
    wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

    ws = wb.active
    
    for cell in ws["{}".format(pam3)]:
        if cell.value is None:
            a = cell.row
            b = 7
            c = ws.cell(row=a, column=b)
            if c.value is not None:
                if fatia(c.value) == "4":
                    cell.value = "TRF"
                    print(cell, cell.value)
            if c.value is not None:
                if fatia(c.value) == "5":
                    cell.value = "TRT"
                    print(cell, cell.value)
            if c.value is not None:
                if fatia(c.value) == "8":
                    cell.value = "TJ"
                    print(cell, cell.value)

    wb.save("importacao_format.xlsx")
