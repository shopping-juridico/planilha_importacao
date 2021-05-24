from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from validacpf import ValidaCpf
from validacnpj import ValidaCnpj

wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

# pf/pj
for cell in ws['S']:
    if(cell.value is not None):
        if cell.value == "Contrário Principal - CPF/CNPJ": #Cliente principal - CPF/CNPJ
            continue
        cpf = ValidaCpf(cell.value)
        cnpj = ValidaCnpj(cell.value)
        if cpf.valida():
            a = 2
            #print('CPF válido')
            #print(cell.value)
        elif cnpj.valida():
            a = 3
            #print('CNPJ válido')
            #print(cell.value)
        else:
            cell.fill = redFill
            print('CPF inválido')
    #else:

wb.save("/home/vitor/projetos/planilha importacao/importacao_format.xlsx")