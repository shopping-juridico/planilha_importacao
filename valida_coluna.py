from openpyxl import load_workbook
from validacpf import ValidaCpf
from validacnpj import ValidaCnpj
from openpyxl.styles import Color, PatternFill

wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

def valida_coluna(pam2):
    
    for cell in ws['{}'.format(pam2)]:
        if(cell.value is not None):
            if cell.value == "Contrário Principal - CPF/CNPJ":
                continue
            if cell.value == "Cliente principal - CPF/CNPJ":
                continue
            cpf = ValidaCpf(cell.value)
            cnpj = ValidaCnpj(cell.value)
            if cpf.valida():
                #a = 2
                print('CPF válido')
                #print(cell.value)
            elif cnpj.valida():
                #a = 3
                print('CNPJ válido')
                #print(cell.value)
            else:
                cell.fill = redFill
                print('CPF inválido')
        #if cell.value is None:

valida_coluna(pam2='P')
valida_coluna(pam2='S')

wb.save("/home/vitor/projetos/planilha_importacao/importacao_format.xlsx")