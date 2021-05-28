from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from method import valida_coluna, classifica_pfpj

wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

valida_coluna(pam2='P')
valida_coluna(pam2='S')
classifica_pfpj(pam1='N')

wb.save("/home/vitor/projetos/planilha_importacao/importacao_format.xlsx")