from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from validacpf import ValidaCpf
from validacnpj import ValidaCnpj

wb = load_workbook('Importação de processos (Aberto) Corrigido.xlsx')

ws = wb.active

redFill = PatternFill(start_color='FFFF0000',
                   end_color='FFFF0000',
                   fill_type='solid')

def classifica_pfpj(pam):

    for cell in ws['{}'.format(pam)]:
        if 'S/A' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'LTDA' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'CIA' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'COMPANH' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'COND' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'ASSOCI' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'SEGUR' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif '-' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)
        elif 'INSTI' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)                  
        elif 'ADVO' in cell.value:
            a = cell.row
            b = 17
            #print(a, b)
            c = ws.cell(row=a, column=b)
            c.value = 'PJ'
            print(c.value)

classifica_pfpj(pam='N')

wb.save("/home/vitor/projetos/planilha_importacao/importacao_format.xlsx")